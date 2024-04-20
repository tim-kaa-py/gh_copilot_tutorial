import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

def load_csv_data(filepath):
    """Load CSV data into a DataFrame."""
    df = pd.read_csv(filepath)
    return df

def calculate_total_loan_cost(df):
    """Calculate the total cost of the loan and add it as a new column."""
    df['TotalLoanCost'] = df['LoanAmount'] + (df['LoanAmount'] * df['InterestRate'] * df['LoanDuration'])
    return df

def identify_insurance_targets(df, interest_rate_threshold, loan_duration_threshold):
    """Identify potential insurance targets based on interest rate and loan duration thresholds."""
    # Create a new column for InsuranceOffer
    df['InsuranceOffer'] = False

    # Iterate over each row in the DataFrame
    for index, row in df.iterrows():
        # Check if interest rate and loan duration meet the thresholds
        if row['InterestRate'] >= interest_rate_threshold and row['LoanDuration'] >= loan_duration_threshold:
            df.at[index, 'InsuranceOffer'] = True

    return df

def export_to_excel(df, filepath):
    """Export the DataFrame to an Excel file."""
    df.to_excel(filepath, index=False)

def format_excel_header(filepath):
    """Format the header of the Excel file with a yellow fill."""
    # Load the Excel file into an openpyxl workbook
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active

    # Define the yellow fill
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Apply the yellow fill to the header row
    for cell in sheet["1:1"]:
        cell.fill = yellow_fill

    # Save the changes to the file
    workbook.save(filepath)

def main():
    """Main function to load data, process it, and export it to an Excel file."""
    # Load the data
    filepath = 'data/input/customer_loan_data.csv'
    df = load_csv_data(filepath)

    # Calculate total loan cost
    df = calculate_total_loan_cost(df)

    # Identify insurance targets
    interest_rate_threshold = 0.05  # 5%
    loan_duration_threshold = 5  # 5 years
    df = identify_insurance_targets(df, interest_rate_threshold, loan_duration_threshold)

    # Export the DataFrame to an Excel file
    output_filepath = 'data/output/processed_loan_data.xlsx'
    export_to_excel(df, output_filepath)

    # Format the header of the Excel file
    format_excel_header(output_filepath)

if __name__ == "__main__":
    main()